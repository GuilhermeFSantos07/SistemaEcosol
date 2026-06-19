# =============================================================================
# relatorios.py
# Tela de relatórios e estatísticas do sistema ECOSOL AM.
#
# Estrutura da tela:
#   1. Dashboard de estatísticas (cards): total de cadastros, com CNPJ, com CPF,
#      homens, mulheres, e distribuição por local de cadastro.
#      → Botão para exportar esse resumo em Excel.
#   2. Exportador customizável: a pessoa filtra por tipo de cadastro, local e
#      intervalo de datas, escolhe quais colunas quer no arquivo, e exporta
#      os cadastros filtrados para Excel.
#
# Segue o mesmo padrão visual de form_ecosol.py, sincronizacao.py e cadastros.py:
#   - Cores lidas do .env via os.getenv()
#   - GroupBox com bordas e título colorido em COR_PRIMARIA
#   - Título principal com linha decorativa inferior
# =============================================================================

import sqlite3                      # Banco de dados local SQLite
import os                           # Leitura das variáveis de ambiente do .env
from datetime import datetime       # Usado para nomear o arquivo exportado com data/hora
from dotenv import load_dotenv      # Carrega o .env para que os.getenv() funcione

from PyQt6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QGridLayout,   # Layouts principais
    QLabel, QPushButton, QComboBox, QDateEdit,         # Widgets de filtro
    QGroupBox, QScrollArea, QCheckBox,                 # Agrupadores, scroll e checkboxes
    QFileDialog, QMessageBox, QFrame                   # Diálogo de salvar arquivo e alertas
)
from PyQt6.QtCore import Qt, QDate    # Constantes de alinhamento e manipulação de datas

# openpyxl é usado para gerar o arquivo .xlsx (precisa estar no requirements.txt)
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# Carrega as variáveis do arquivo .env
load_dotenv()

# =============================================================================
# PALETA DE CORES — lidas do .env, idênticas aos demais arquivos do sistema
# =============================================================================
COR_PRIMARIA       = os.getenv("COR_PRIMARIA",       "#003366")  # Azul escuro — títulos, bordas
COR_PRIMARIA_HOVER = os.getenv("COR_PRIMARIA_HOVER", "#004080")  # Azul hover dos botões
COR_SECUNDARIA     = os.getenv("COR_SECUNDARIA",     "#17a2b8")  # Azul ciano — botões de exportar
COR_FUNDO_CLARO    = os.getenv("COR_FUNDO_CLARO",   "#f8f9fa")  # Cinza claro — fundo dos cards
COR_TEXTO_ESCURO   = os.getenv("COR_TEXTO_ESCURO",   "#212529")  # Quase preto — texto dos labels
COR_TEXTO_CLARO    = os.getenv("COR_TEXTO_CLARO",    "#ffffff")  # Branco — texto sobre fundo colorido
COR_BORDA          = os.getenv("COR_BORDA",          "#ced4da")  # Cinza suave — bordas dos cards/inputs
COR_SALVAR         = "#28a745"   # Verde fixo — botão de exportar cadastros (mesmo padrão do form_ecosol)
COR_SALVAR_HOVER   = "#218838"   # Verde escuro — hover do botão de exportar

# =============================================================================
# Lista de TODAS as colunas exportáveis no modo customizável.
# Cada tupla: (nome_coluna_banco, rótulo_amigável_exibido_no_checkbox)
# Usada tanto para montar os checkboxes quanto para montar o SELECT do banco.
# =============================================================================
COLUNAS_EXPORTAVEIS = [
    ("id",                          "ID do Cadastro"),
    ("tipo_cadastro",               "Tipo de Cadastro"),
    ("razao_social_nome",           "Nome / Razão Social"),
    ("representante_legal",         "Representante Legal"),
    ("cpf",                         "CPF"),
    ("cnpj",                        "CNPJ"),
    ("rg",                          "RG"),
    ("telefone",                    "Telefone"),
    ("email",                       "E-mail"),
    ("endereco",                    "Endereço"),
    ("cep",                         "CEP"),
    ("sexo",                        "Sexo"),
    ("cor_raca",                    "Cor/Raça"),
    ("forma_organizacao_ecosol",    "Forma Org. ECOSOL"),
    ("forma_organizacao_emp",       "Forma Org. Empreendimento"),
    ("segmento_empreendimento",     "Segmento do Empreendimento"),
    ("materia_prima",               "Matéria-Prima"),
    ("local_producao",              "Local de Produção"),
    ("onde_comercializa",           "Onde Comercializa"),
    ("beneficiarios_diretos_m",     "Beneficiários Diretos (M)"),
    ("beneficiarios_diretos_f",     "Beneficiários Diretos (F)"),
    ("beneficiarios_indiretos_m",   "Beneficiários Indiretos (M)"),
    ("beneficiarios_indiretos_f",   "Beneficiários Indiretos (F)"),
    ("maquina_cartao",              "Possui Máquina de Cartão"),
    ("pix",                         "Possui PIX"),
    ("classificacao_social",       "Classificação Social"),
    ("motivo_criacao",              "Motivo de Criação"),
    ("formas_comercializacao",      "Formas de Comercialização"),
    ("produtos_comercializados",    "Produtos Comercializados"),
    ("pagam_taxa",                  "Pagam Taxa/Contribuição"),
    ("forma_contribuicao",          "Forma de Contribuição"),
    ("renda_preponderante",         "Renda Preponderante"),
    ("para_quem_comercializa",      "Para Quem Comercializa"),
    ("dificuldade_comercializacao", "Dificuldade de Comercialização"),
    ("responsavel_vendas",          "Responsável pelas Vendas"),
    ("obs",                         "Observações"),
    ("local_cadastro",              "Local do Cadastro"),
    ("data_formulario",             "Data do Formulário"),
    ("data_cadastro",               "Data/Hora do Cadastro"),
]

# Colunas marcadas por padrão ao abrir a tela (as mais usadas no dia a dia)
COLUNAS_PADRAO_MARCADAS = {
    "razao_social_nome", "representante_legal", "cpf", "cnpj",
    "telefone", "email", "local_cadastro", "data_cadastro",
}


# =============================================================================
# TelaRelatorios
# Widget principal da tela de relatórios e estatísticas.
# =============================================================================
class TelaRelatorios(QWidget):
    def __init__(self):
        super().__init__()

        # Lista que guardará os QCheckBox de colunas para leitura posterior
        self.checks_colunas = []

        # Dicionário que guardará os QLabel dos cards de estatística para atualização
        self.labels_estatisticas = {}

        self.configurar_ui()        # Monta toda a interface da tela
        self.atualizar_estatisticas()  # Carrega os números na primeira abertura
        self.popular_filtro_tipo()     # Popula o combo de tipos de cadastro existentes
        self.popular_filtro_local()    # Popula o combo de locais de cadastro existentes

    # =========================================================================
    # configurar_ui
    # Monta a estrutura completa da tela dentro de um QScrollArea, já que o
    # conteúdo (dashboard + exportador) pode ser maior que a altura da janela.
    # =========================================================================
    def configurar_ui(self):
        # Layout raiz da tela — contém apenas a área de rolagem
        layout_raiz = QVBoxLayout(self)
        layout_raiz.setContentsMargins(0, 0, 0, 0)
        layout_raiz.setSpacing(0)

        # ----- ÁREA DE ROLAGEM -----
        # Permite que o conteúdo (estatísticas + exportador) role verticalmente
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)                      # Conteúdo interno se redimensiona
        scroll.setStyleSheet("QScrollArea { border: none; background-color: transparent; }")

        # Widget interno que recebe todo o conteúdo da tela
        conteudo = QWidget()
        layout_principal = QVBoxLayout(conteudo)
        layout_principal.setContentsMargins(30, 20, 30, 20)  # Margens internas da tela
        layout_principal.setSpacing(18)                        # Espaço entre os blocos

        # ----- CABEÇALHO: título com linha decorativa -----
        titulo = QLabel("RELATÓRIOS E ESTATÍSTICAS")
        titulo.setStyleSheet(
            f"font-size: 22px; font-weight: bold; "
            f"color: {COR_PRIMARIA}; "
            f"border-bottom: 2px solid {COR_PRIMARIA}; "
            f"padding-bottom: 5px;"
        )
        layout_principal.addWidget(titulo)

        # Subtítulo descritivo
        subtitulo = QLabel("Visualize estatísticas gerais e exporte dados para Excel")
        subtitulo.setStyleSheet("color: #6c757d; font-size: 12px; margin-top: -4px;")
        layout_principal.addWidget(subtitulo)

        # =================================================================
        # SEÇÃO 1: DASHBOARD DE ESTATÍSTICAS
        # =================================================================
        layout_principal.addWidget(self._criar_secao_dashboard())

        # ----- Separador visual entre as duas seções -----
        separador = QFrame()
        separador.setFrameShape(QFrame.Shape.HLine)         # Linha horizontal
        separador.setStyleSheet(f"background-color: {COR_BORDA}; max-height: 1px; border: none;")
        layout_principal.addWidget(separador)

        # =================================================================
        # SEÇÃO 2: EXPORTADOR CUSTOMIZÁVEL DE CADASTROS
        # =================================================================
        layout_principal.addWidget(self._criar_secao_exportador())

        layout_principal.addStretch()  # Empurra tudo para cima quando há espaço sobrando

        # Finaliza: coloca o conteúdo dentro do scroll e adiciona à tela
        scroll.setWidget(conteudo)
        layout_raiz.addWidget(scroll)

    # =========================================================================
    # _criar_secao_dashboard
    # Monta o GroupBox com os cards de estatística e o botão de exportar resumo.
    # Retorna o QGroupBox pronto para ser adicionado ao layout principal.
    # =========================================================================
    def _criar_secao_dashboard(self):
        grupo = QGroupBox("Resumo Geral dos Cadastros")
        grupo.setStyleSheet(f"""
            QGroupBox {{
                font-weight: bold;
                font-size: 14px;
                border: 1px solid {COR_BORDA};
                border-radius: 6px;
                margin-top: 12px;
                padding: 20px 16px 16px 16px;
                background-color: {COR_TEXTO_CLARO};
            }}
            QGroupBox::title {{
                subcontrol-origin: margin;
                subcontrol-position: top left;
                left: 15px;
                padding: 0 5px;
                color: {COR_PRIMARIA};
            }}
        """)

        layout_grupo = QVBoxLayout()
        layout_grupo.setSpacing(16)

        # ----- Grid de cards de estatística -----
        # Cada card é um pequeno QFrame com um número grande e um rótulo
        grid_cards = QGridLayout()
        grid_cards.setSpacing(14)

        # Lista de cards a criar: (chave_interna, rótulo_exibido, cor_destaque)
        # A "chave_interna" é usada como índice no dicionário self.labels_estatisticas
        definicao_cards = [
            ("total",       "Total de Cadastros",     COR_PRIMARIA),
            ("com_cnpj",    "Com CNPJ",                COR_SECUNDARIA),
            ("com_cpf",     "Com CPF",                 COR_SECUNDARIA),
            ("homens",      "Sexo Masculino",          "#0d6efd"),
            ("mulheres",    "Sexo Feminino",           "#d63384"),
        ]

        # Cria os 5 cards principais em uma grade de até 5 colunas
        for indice, (chave, rotulo, cor) in enumerate(definicao_cards):
            card = self._criar_card_estatistica(chave, rotulo, cor)
            grid_cards.addWidget(card, 0, indice)  # Todos na linha 0, uma coluna cada

        layout_grupo.addLayout(grid_cards)

        # ----- Bloco: distribuição por Local de Cadastro -----
        # Diferente dos cards fixos acima, este bloco é dinâmico (varia conforme os locais no banco)
        lbl_local_titulo = QLabel("Distribuição por Local de Cadastro:")
        lbl_local_titulo.setStyleSheet(f"font-size: 13px; font-weight: bold; color: {COR_TEXTO_ESCURO}; margin-top: 8px;")
        layout_grupo.addWidget(lbl_local_titulo)

        # Container onde os mini-cards de cada local serão inseridos dinamicamente
        self.layout_locais = QHBoxLayout()
        self.layout_locais.setSpacing(10)
        self.widget_locais = QWidget()  # Necessário para poder limpar/recriar o conteúdo depois
        self.widget_locais.setLayout(self.layout_locais)
        layout_grupo.addWidget(self.widget_locais)

        # ----- Botão: Exportar Resumo em Excel -----
        layout_btn_resumo = QHBoxLayout()
        layout_btn_resumo.addStretch()  # Empurra o botão para a direita

        btn_exportar_resumo = QPushButton("📊  Exportar Resumo em Excel")
        btn_exportar_resumo.setFixedWidth(260)
        btn_exportar_resumo.setStyleSheet(f"""
            QPushButton {{
                background-color: {COR_SECUNDARIA};
                color: {COR_TEXTO_CLARO};
                padding: 10px 18px;
                font-size: 13px;
                font-weight: bold;
                border-radius: 4px;
                border: none;
            }}
            QPushButton:hover {{ background-color: {COR_PRIMARIA}; }}
        """)
        btn_exportar_resumo.clicked.connect(self.exportar_resumo_excel)  # Conecta ao método de exportação
        layout_btn_resumo.addWidget(btn_exportar_resumo)

        layout_grupo.addLayout(layout_btn_resumo)

        grupo.setLayout(layout_grupo)
        return grupo

    # =========================================================================
    # _criar_card_estatistica
    # Cria um único "card" visual (QFrame) com um número grande e um rótulo.
    # Guarda o QLabel do número no dicionário self.labels_estatisticas para que
    # atualizar_estatisticas() possa alterar o valor exibido depois.
    #
    # Parâmetros:
    #   chave  → identificador interno (usado como chave do dicionário)
    #   rotulo → texto descritivo exibido abaixo do número
    #   cor    → cor de destaque do número e da borda esquerda do card
    # =========================================================================
    def _criar_card_estatistica(self, chave, rotulo, cor):
        card = QFrame()
        card.setStyleSheet(f"""
            QFrame {{
                background-color: {COR_FUNDO_CLARO};
                border: 1px solid {COR_BORDA};
                border-left: 4px solid {cor};
                border-radius: 6px;
            }}
        """)
        card.setMinimumWidth(150)   # Largura mínima para não ficar apertado
        card.setMinimumHeight(80)   # Altura mínima uniforme entre os cards

        layout_card = QVBoxLayout(card)
        layout_card.setContentsMargins(14, 10, 14, 10)
        layout_card.setSpacing(2)

        # Número grande — valor inicial "0", atualizado depois por atualizar_estatisticas()
        lbl_numero = QLabel("0")
        lbl_numero.setStyleSheet(f"font-size: 26px; font-weight: bold; color: {cor}; background: transparent;")
        layout_card.addWidget(lbl_numero)

        # Rótulo descritivo abaixo do número
        lbl_rotulo = QLabel(rotulo)
        lbl_rotulo.setStyleSheet(f"font-size: 11px; color: #6c757d; background: transparent;")
        lbl_rotulo.setWordWrap(True)  # Permite quebrar em duas linhas se o texto for longo
        layout_card.addWidget(lbl_rotulo)

        # Guarda a referência ao label do número para poder atualizá-lo depois
        self.labels_estatisticas[chave] = lbl_numero

        return card

    # =========================================================================
    # atualizar_estatisticas
    # Consulta o banco SQLite e atualiza todos os números exibidos nos cards.
    # Também recria os mini-cards de distribuição por local de cadastro.
    # Chamado na abertura da tela e pode ser chamado novamente se necessário.
    # =========================================================================
    def atualizar_estatisticas(self):
        conn   = sqlite3.connect('ecosol_local.db')  # Abre conexão com o banco local
        cursor = conn.cursor()

        # ----- Total de cadastros -----
        cursor.execute("SELECT COUNT(*) FROM cadastros_ecosol")
        total = cursor.fetchone()[0]  # fetchone retorna uma tupla; [0] pega o primeiro valor

        # ----- Quantidade com CNPJ preenchido (não nulo e não vazio) -----
        cursor.execute("SELECT COUNT(*) FROM cadastros_ecosol WHERE cnpj IS NOT NULL AND cnpj != ''")
        com_cnpj = cursor.fetchone()[0]

        # ----- Quantidade com CPF preenchido -----
        cursor.execute("SELECT COUNT(*) FROM cadastros_ecosol WHERE cpf IS NOT NULL AND cpf != ''")
        com_cpf = cursor.fetchone()[0]

        # ----- Quantidade do sexo Masculino -----
        # Usa LIKE para cobrir variações de capitalização ("Masculino", "masculino")
        cursor.execute("SELECT COUNT(*) FROM cadastros_ecosol WHERE sexo LIKE 'Masculino%'")
        homens = cursor.fetchone()[0]

        # ----- Quantidade do sexo Feminino -----
        cursor.execute("SELECT COUNT(*) FROM cadastros_ecosol WHERE sexo LIKE 'Feminino%'")
        mulheres = cursor.fetchone()[0]

        # ----- Distribuição por local de cadastro (agrupado, contagem decrescente) -----
        cursor.execute("""
            SELECT local_cadastro, COUNT(*) as qtd
            FROM cadastros_ecosol
            WHERE local_cadastro IS NOT NULL AND local_cadastro != ''
            GROUP BY local_cadastro
            ORDER BY qtd DESC
        """)
        distribuicao_local = cursor.fetchall()  # Lista de tuplas (local, quantidade)

        conn.close()  # Fecha a conexão imediatamente após todas as consultas

        # ----- Atualiza o texto dos cards principais -----
        self.labels_estatisticas["total"].setText(str(total))
        self.labels_estatisticas["com_cnpj"].setText(str(com_cnpj))
        self.labels_estatisticas["com_cpf"].setText(str(com_cpf))
        self.labels_estatisticas["homens"].setText(str(homens))
        self.labels_estatisticas["mulheres"].setText(str(mulheres))

        # ----- Recria os mini-cards de distribuição por local -----
        self._atualizar_cards_locais(distribuicao_local)

    # =========================================================================
    # _atualizar_cards_locais
    # Remove os mini-cards de local antigos e cria novos com base na consulta
    # mais recente. Necessário porque a quantidade de locais é dinâmica
    # (depende de quais locais existem no banco no momento).
    #
    # Parâmetros:
    #   distribuicao → lista de tuplas (nome_local, quantidade)
    # =========================================================================
    def _atualizar_cards_locais(self, distribuicao):
        # Remove todos os widgets antigos do layout antes de adicionar os novos
        while self.layout_locais.count():
            item = self.layout_locais.takeAt(0)   # Remove o item da posição 0
            widget = item.widget()
            if widget:
                widget.deleteLater()  # Agenda a destruição do widget antigo

        if not distribuicao:
            # Nenhum cadastro com local preenchido ainda
            lbl_vazio = QLabel("Nenhum dado de local disponível ainda.")
            lbl_vazio.setStyleSheet("color: #adb5bd; font-style: italic; font-size: 12px;")
            self.layout_locais.addWidget(lbl_vazio)
            return

        # Cria um mini-card simples para cada local encontrado
        for local, quantidade in distribuicao:
            mini_card = QFrame()
            mini_card.setStyleSheet(f"""
                QFrame {{
                    background-color: {COR_FUNDO_CLARO};
                    border: 1px solid {COR_BORDA};
                    border-radius: 14px;
                    padding: 2px;
                }}
            """)
            layout_mini = QHBoxLayout(mini_card)
            layout_mini.setContentsMargins(12, 6, 12, 6)
            layout_mini.setSpacing(6)

            # Texto: "Manaus: 12"
            lbl_texto = QLabel(f"{local}:")
            lbl_texto.setStyleSheet(f"font-size: 12px; color: {COR_TEXTO_ESCURO}; background: transparent;")
            layout_mini.addWidget(lbl_texto)

            lbl_qtd = QLabel(str(quantidade))
            lbl_qtd.setStyleSheet(f"font-size: 12px; font-weight: bold; color: {COR_PRIMARIA}; background: transparent;")
            layout_mini.addWidget(lbl_qtd)

            self.layout_locais.addWidget(mini_card)

        self.layout_locais.addStretch()  # Empurra os mini-cards para a esquerda

    # =========================================================================
    # _criar_secao_exportador
    # Monta o GroupBox do exportador customizável: filtros + checkboxes de
    # colunas + botão de exportar. Retorna o QGroupBox pronto.
    # =========================================================================
    def _criar_secao_exportador(self):
        grupo = QGroupBox("Exportar Cadastros para Excel (Personalizado)")
        grupo.setStyleSheet(f"""
            QGroupBox {{
                font-weight: bold;
                font-size: 14px;
                border: 1px solid {COR_BORDA};
                border-radius: 6px;
                margin-top: 12px;
                padding: 20px 16px 16px 16px;
                background-color: {COR_TEXTO_CLARO};
            }}
            QGroupBox::title {{
                subcontrol-origin: margin;
                subcontrol-position: top left;
                left: 15px;
                padding: 0 5px;
                color: {COR_PRIMARIA};
            }}
            QComboBox, QDateEdit {{
                border: 1px solid {COR_BORDA};
                border-radius: 4px;
                padding: 6px;
                background-color: {COR_TEXTO_CLARO};
                min-height: 26px;
            }}
            QComboBox:focus, QDateEdit:focus {{
                border: 1px solid {COR_PRIMARIA};
            }}
            QLabel {{
                font-size: 13px;
                color: {COR_TEXTO_ESCURO};
            }}
        """)

        layout_grupo = QVBoxLayout()
        layout_grupo.setSpacing(14)

        # =================================================================
        # SUB-BLOCO: FILTROS (tipo de cadastro, local, intervalo de datas)
        # =================================================================
        lbl_filtros = QLabel("Filtros (deixe em branco para não filtrar por aquele campo):")
        lbl_filtros.setStyleSheet(f"font-size: 12px; font-weight: bold; color: {COR_TEXTO_ESCURO};")
        layout_grupo.addWidget(lbl_filtros)

        grid_filtros = QGridLayout()
        grid_filtros.setSpacing(10)
        grid_filtros.setColumnStretch(1, 1)  # Coluna de inputs elástica
        grid_filtros.setColumnStretch(3, 1)

        # ----- Filtro: Tipo de Cadastro -----
        self.filtro_tipo = QComboBox()
        self.filtro_tipo.addItem("Todos")  # Opção inicial = sem filtro
        grid_filtros.addWidget(QLabel("Tipo de Cadastro:"), 0, 0)
        grid_filtros.addWidget(self.filtro_tipo, 0, 1)

        # ----- Filtro: Local de Cadastro -----
        self.filtro_local = QComboBox()
        self.filtro_local.addItem("Todos")  # Opção inicial = sem filtro
        grid_filtros.addWidget(QLabel("Local de Cadastro:"), 0, 2)
        grid_filtros.addWidget(self.filtro_local, 0, 3)

        # ----- Filtro: Data Inicial -----
        self.filtro_data_inicio = QDateEdit()
        self.filtro_data_inicio.setCalendarPopup(True)             # Abre calendário ao clicar
        self.filtro_data_inicio.setDisplayFormat("dd/MM/yyyy")     # Formato visual brasileiro
        self.filtro_data_inicio.setDate(QDate(2020, 1, 1))         # Data inicial padrão bem antiga
        grid_filtros.addWidget(QLabel("Data Inicial:"), 1, 0)
        grid_filtros.addWidget(self.filtro_data_inicio, 1, 1)

        # ----- Filtro: Data Final -----
        self.filtro_data_fim = QDateEdit()
        self.filtro_data_fim.setCalendarPopup(True)
        self.filtro_data_fim.setDisplayFormat("dd/MM/yyyy")
        self.filtro_data_fim.setDate(QDate.currentDate())          # Data final padrão = hoje
        grid_filtros.addWidget(QLabel("Data Final:"), 1, 2)
        grid_filtros.addWidget(self.filtro_data_fim, 1, 3)

        # ----- Checkbox: Habilitar filtro de data -----
        # Por padrão o filtro de data fica desabilitado (exporta todas as datas)
        self.check_usar_filtro_data = QCheckBox("Filtrar por intervalo de datas")
        self.check_usar_filtro_data.toggled.connect(self.filtro_data_inicio.setEnabled)
        self.check_usar_filtro_data.toggled.connect(self.filtro_data_fim.setEnabled)
        self.filtro_data_inicio.setEnabled(False)  # Começa desabilitado
        self.filtro_data_fim.setEnabled(False)
        grid_filtros.addWidget(self.check_usar_filtro_data, 2, 0, 1, 2)

        layout_grupo.addLayout(grid_filtros)

        # =================================================================
        # SUB-BLOCO: SELEÇÃO DE COLUNAS (checkboxes em grid de 4 colunas)
        # =================================================================
        lbl_colunas = QLabel("Selecione as colunas a incluir no arquivo Excel:")
        lbl_colunas.setStyleSheet(f"font-size: 12px; font-weight: bold; color: {COR_TEXTO_ESCURO}; margin-top: 6px;")
        layout_grupo.addWidget(lbl_colunas)

        # ----- Botões de atalho: Marcar Todas / Desmarcar Todas -----
        layout_atalhos = QHBoxLayout()
        btn_marcar_todas = QPushButton("Marcar Todas")
        btn_marcar_todas.setFixedWidth(130)
        btn_marcar_todas.setStyleSheet(f"""
            QPushButton {{
                background-color: {COR_PRIMARIA}; color: {COR_TEXTO_CLARO};
                padding: 5px 10px; font-size: 11px; border-radius: 4px; border: none;
            }}
            QPushButton:hover {{ background-color: {COR_PRIMARIA_HOVER}; }}
        """)
        btn_marcar_todas.clicked.connect(lambda: self._marcar_todas_colunas(True))  # Marca todos os checkboxes
        layout_atalhos.addWidget(btn_marcar_todas)

        btn_desmarcar_todas = QPushButton("Desmarcar Todas")
        btn_desmarcar_todas.setFixedWidth(130)
        btn_desmarcar_todas.setStyleSheet(f"""
            QPushButton {{
                background-color: #6c757d; color: {COR_TEXTO_CLARO};
                padding: 5px 10px; font-size: 11px; border-radius: 4px; border: none;
            }}
            QPushButton:hover {{ background-color: #5a6268; }}
        """)
        btn_desmarcar_todas.clicked.connect(lambda: self._marcar_todas_colunas(False))  # Desmarca todos
        layout_atalhos.addWidget(btn_desmarcar_todas)

        layout_atalhos.addStretch()  # Empurra os botões de atalho para a esquerda
        layout_grupo.addLayout(layout_atalhos)

        # ----- Grid de checkboxes de colunas (4 por linha) -----
        widget_checkboxes = QWidget()
        widget_checkboxes.setStyleSheet(f"background-color: {COR_FUNDO_CLARO}; border-radius: 6px;")
        grid_checkboxes = QGridLayout(widget_checkboxes)
        grid_checkboxes.setContentsMargins(10, 10, 10, 10)
        grid_checkboxes.setSpacing(8)

        colunas_por_linha = 4  # Máximo de checkboxes por linha do grid
        col = 0
        row = 0

        for nome_coluna, rotulo in COLUNAS_EXPORTAVEIS:
            cb = QCheckBox(rotulo)
            cb.setStyleSheet(f"font-size: 12px; color: {COR_TEXTO_ESCURO}; background: transparent;")

            # Marca por padrão as colunas mais usadas (definidas em COLUNAS_PADRAO_MARCADAS)
            cb.setChecked(nome_coluna in COLUNAS_PADRAO_MARCADAS)

            # Guarda o nome real da coluna do banco como propriedade do checkbox,
            # para que possamos recuperá-lo depois sem depender da ordem da lista
            cb.setProperty("nome_coluna", nome_coluna)

            self.checks_colunas.append(cb)  # Guarda referência para uso posterior
            grid_checkboxes.addWidget(cb, row, col)

            col += 1
            if col >= colunas_por_linha:  # Quebra de linha a cada 4 colunas
                col = 0
                row += 1

        layout_grupo.addWidget(widget_checkboxes)

        # =================================================================
        # BOTÃO: EXPORTAR CADASTROS FILTRADOS
        # =================================================================
        layout_btn_export = QHBoxLayout()
        layout_btn_export.addStretch()  # Empurra o botão para a direita

        btn_exportar = QPushButton("📥  Exportar Cadastros Filtrados em Excel")
        btn_exportar.setFixedWidth(320)
        btn_exportar.setStyleSheet(f"""
            QPushButton {{
                background-color: {COR_SALVAR};
                color: {COR_TEXTO_CLARO};
                padding: 10px 18px;
                font-size: 13px;
                font-weight: bold;
                border-radius: 4px;
                border: none;
            }}
            QPushButton:hover {{ background-color: {COR_SALVAR_HOVER}; }}
        """)
        btn_exportar.clicked.connect(self.exportar_cadastros_excel)  # Conecta ao método de exportação
        layout_btn_export.addWidget(btn_exportar)

        layout_grupo.addLayout(layout_btn_export)

        grupo.setLayout(layout_grupo)
        return grupo

    # =========================================================================
    # _marcar_todas_colunas
    # Marca ou desmarca todos os checkboxes de colunas de uma vez.
    # Chamado pelos botões de atalho "Marcar Todas" / "Desmarcar Todas".
    #
    # Parâmetros:
    #   marcar → True para marcar todos, False para desmarcar todos
    # =========================================================================
    def _marcar_todas_colunas(self, marcar: bool):
        for cb in self.checks_colunas:
            cb.setChecked(marcar)  # Aplica o mesmo estado a todos os checkboxes

    # =========================================================================
    # popular_filtro_tipo
    # Consulta os tipos de cadastro distintos já usados no banco e os adiciona
    # ao combo de filtro. Evita que o usuário digite manualmente o tipo exato.
    # =========================================================================
    def popular_filtro_tipo(self):
        conn   = sqlite3.connect('ecosol_local.db')
        cursor = conn.cursor()

        # SELECT DISTINCT retorna cada valor de tipo_cadastro apenas uma vez
        cursor.execute("""
            SELECT DISTINCT tipo_cadastro FROM cadastros_ecosol
            WHERE tipo_cadastro IS NOT NULL AND tipo_cadastro != ''
            ORDER BY tipo_cadastro
        """)
        tipos = cursor.fetchall()  # Lista de tuplas de 1 elemento cada
        conn.close()

        for (tipo,) in tipos:  # Desempacota cada tupla de 1 elemento
            self.filtro_tipo.addItem(tipo)  # Adiciona cada tipo encontrado ao combo

    # =========================================================================
    # popular_filtro_local
    # Mesma lógica de popular_filtro_tipo, mas para a coluna local_cadastro.
    # =========================================================================
    def popular_filtro_local(self):
        conn   = sqlite3.connect('ecosol_local.db')
        cursor = conn.cursor()

        cursor.execute("""
            SELECT DISTINCT local_cadastro FROM cadastros_ecosol
            WHERE local_cadastro IS NOT NULL AND local_cadastro != ''
            ORDER BY local_cadastro
        """)
        locais = cursor.fetchall()
        conn.close()

        for (local,) in locais:
            self.filtro_local.addItem(local)  # Adiciona cada local encontrado ao combo

    # =========================================================================
    # exportar_resumo_excel
    # Gera um arquivo .xlsx com o resumo estatístico (mesmos números dos cards)
    # e abre um diálogo para o usuário escolher onde salvar.
    # =========================================================================
    def exportar_resumo_excel(self):
        # Abre o diálogo "Salvar Como" com nome de arquivo sugerido já com data/hora
        nome_sugerido = f"Resumo_Estatisticas_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        caminho, _ = QFileDialog.getSaveFileName(
            self, "Salvar Resumo em Excel", nome_sugerido, "Planilha Excel (*.xlsx)"
        )

        if not caminho:
            return  # Usuário cancelou o diálogo — não faz nada

        try:
            # ----- Cria a planilha em branco -----
            wb = Workbook()
            sheet = wb.active
            sheet.title = "Resumo Estatísticas"

            # ----- Cabeçalho do relatório -----
            sheet["A1"] = "RELATÓRIO DE ESTATÍSTICAS — SISTEMA ECOSOL AM"
            sheet["A1"].font = Font(bold=True, size=14, color="003366")  # Azul institucional
            sheet.merge_cells("A1:B1")  # Mescla para o título ocupar 2 colunas

            sheet["A2"] = f"Gerado em: {datetime.now().strftime('%d/%m/%Y às %H:%M')}"
            sheet["A2"].font = Font(italic=True, size=9, color="6c757d")
            sheet.merge_cells("A2:B2")

            # ----- Cabeçalho da tabela de dados (linha 4) -----
            sheet["A4"] = "Indicador"
            sheet["B4"] = "Valor"
            for celula in ("A4", "B4"):
                sheet[celula].font = Font(bold=True, color="FFFFFF")
                sheet[celula].fill = PatternFill("solid", start_color="003366")  # Fundo azul institucional
                sheet[celula].alignment = Alignment(horizontal="center")

            # ----- Linhas de dados — lidas diretamente dos labels já calculados -----
            linhas_dados = [
                ("Total de Cadastros",  self.labels_estatisticas["total"].text()),
                ("Com CNPJ",            self.labels_estatisticas["com_cnpj"].text()),
                ("Com CPF",             self.labels_estatisticas["com_cpf"].text()),
                ("Sexo Masculino",      self.labels_estatisticas["homens"].text()),
                ("Sexo Feminino",       self.labels_estatisticas["mulheres"].text()),
            ]

            linha_atual = 5  # Primeira linha de dados (depois do cabeçalho na linha 4)
            for rotulo, valor in linhas_dados:
                sheet.cell(row=linha_atual, column=1, value=rotulo)
                sheet.cell(row=linha_atual, column=2, value=int(valor))  # Converte para número
                linha_atual += 1

            # ----- Bloco adicional: distribuição por local de cadastro -----
            linha_atual += 1  # Linha em branco antes do próximo bloco
            sheet.cell(row=linha_atual, column=1, value="Distribuição por Local de Cadastro")
            sheet.cell(row=linha_atual, column=1).font = Font(bold=True, color="003366")
            linha_atual += 1

            # Consulta novamente o banco para obter a distribuição por local
            conn   = sqlite3.connect('ecosol_local.db')
            cursor = conn.cursor()
            cursor.execute("""
                SELECT local_cadastro, COUNT(*) FROM cadastros_ecosol
                WHERE local_cadastro IS NOT NULL AND local_cadastro != ''
                GROUP BY local_cadastro ORDER BY COUNT(*) DESC
            """)
            distribuicao = cursor.fetchall()
            conn.close()

            for local, quantidade in distribuicao:
                sheet.cell(row=linha_atual, column=1, value=local)
                sheet.cell(row=linha_atual, column=2, value=quantidade)
                linha_atual += 1

            # ----- Ajusta a largura das colunas para melhor leitura -----
            sheet.column_dimensions["A"].width = 32
            sheet.column_dimensions["B"].width = 14

            wb.save(caminho)  # Salva o arquivo no caminho escolhido pelo usuário

            QMessageBox.information(self, "Sucesso", f"Resumo exportado com sucesso!\n\n{caminho}")

        except Exception as e:
            # Captura qualquer erro de escrita (permissão negada, disco cheio, etc.)
            QMessageBox.critical(self, "Erro", f"Falha ao exportar o resumo:\n{str(e)}")

    # =========================================================================
    # exportar_cadastros_excel
    # Lê os filtros e as colunas selecionadas pelo usuário, monta a consulta
    # SQL dinamicamente, e gera um arquivo .xlsx com os cadastros filtrados.
    # =========================================================================
    def exportar_cadastros_excel(self):
        # ----- Recupera as colunas marcadas pelo usuário -----
        colunas_selecionadas = [
            cb.property("nome_coluna")          # Nome real da coluna no banco
            for cb in self.checks_colunas
            if cb.isChecked()                    # Apenas os checkboxes marcados
        ]

        if not colunas_selecionadas:
            # Sem nenhuma coluna marcada não há o que exportar
            QMessageBox.warning(self, "Aviso", "Selecione ao menos uma coluna para exportar!")
            return

        # ----- Monta a cláusula WHERE dinamicamente com base nos filtros -----
        condicoes  = []   # Lista de trechos SQL tipo "coluna = ?"
        parametros = []   # Lista de valores correspondentes aos "?" acima

        # Filtro de Tipo de Cadastro (só aplica se não estiver em "Todos")
        if self.filtro_tipo.currentText() != "Todos":
            condicoes.append("tipo_cadastro = ?")
            parametros.append(self.filtro_tipo.currentText())

        # Filtro de Local de Cadastro (só aplica se não estiver em "Todos")
        if self.filtro_local.currentText() != "Todos":
            condicoes.append("local_cadastro = ?")
            parametros.append(self.filtro_local.currentText())

        # Filtro de intervalo de datas (só aplica se o checkbox estiver marcado)
        if self.check_usar_filtro_data.isChecked():
            data_ini = self.filtro_data_inicio.date().toString("yyyy-MM-dd")  # Formato ISO para comparação
            data_fim = self.filtro_data_fim.date().toString("yyyy-MM-dd")
            # data_cadastro é armazenada como "YYYY-MM-DD HH:MM:SS" — comparamos pela parte da data
            condicoes.append("date(data_cadastro) BETWEEN ? AND ?")
            parametros.append(data_ini)
            parametros.append(data_fim)

        # Monta a query final: SELECT colunas FROM tabela [WHERE condições]
        colunas_sql = ", ".join(colunas_selecionadas)  # Ex: "razao_social_nome, cpf, telefone"
        query = f"SELECT {colunas_sql} FROM cadastros_ecosol"

        if condicoes:
            # Junta todas as condições com "AND" — todos os filtros marcados devem ser satisfeitos
            query += " WHERE " + " AND ".join(condicoes)

        query += " ORDER BY data_cadastro DESC"  # Mais recentes primeiro

        # ----- Executa a consulta no banco -----
        conn   = sqlite3.connect('ecosol_local.db')
        cursor = conn.cursor()
        cursor.execute(query, parametros)  # Passa os parâmetros de forma segura (evita SQL injection)
        resultados = cursor.fetchall()
        conn.close()

        if not resultados:
            # Nenhum cadastro corresponde aos filtros aplicados
            QMessageBox.information(self, "Sem Resultados", "Nenhum cadastro encontrado com os filtros selecionados.")
            return

        # ----- Abre o diálogo para escolher onde salvar o arquivo -----
        nome_sugerido = f"Cadastros_Ecosol_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        caminho, _ = QFileDialog.getSaveFileName(
            self, "Salvar Cadastros em Excel", nome_sugerido, "Planilha Excel (*.xlsx)"
        )

        if not caminho:
            return  # Usuário cancelou o diálogo

        try:
            # ----- Cria a planilha e escreve o cabeçalho -----
            wb = Workbook()
            sheet = wb.active
            sheet.title = "Cadastros ECOSOL"

            # Monta os rótulos amigáveis na mesma ordem das colunas selecionadas
            # Cria um dicionário rápido: nome_coluna_banco → rótulo amigável
            mapa_rotulos = dict(COLUNAS_EXPORTAVEIS)
            cabecalhos = [mapa_rotulos.get(col, col) for col in colunas_selecionadas]

            # Escreve a linha de cabeçalho na primeira linha da planilha
            for col_idx, cabecalho in enumerate(cabecalhos, start=1):  # Excel começa em 1, não 0
                celula = sheet.cell(row=1, column=col_idx, value=cabecalho)
                celula.font = Font(bold=True, color="FFFFFF")
                celula.fill = PatternFill("solid", start_color="003366")  # Fundo azul institucional
                celula.alignment = Alignment(horizontal="center")

            # ----- Escreve os dados linha por linha a partir da linha 2 -----
            for linha_idx, linha_dados in enumerate(resultados, start=2):  # Linha 2 = primeira linha de dados
                for col_idx, valor in enumerate(linha_dados, start=1):
                    sheet.cell(row=linha_idx, column=col_idx, value=valor)

            # ----- Ajusta a largura de cada coluna com base no maior conteúdo -----
            for col_idx, cabecalho in enumerate(cabecalhos, start=1):
                letra_coluna = get_column_letter(col_idx)  # Ex: 1 → "A", 2 → "B"
                # Largura proporcional ao tamanho do cabeçalho, com limite mínimo e máximo
                largura = max(len(cabecalho) + 2, 14)
                sheet.column_dimensions[letra_coluna].width = min(largura, 40)

            # ----- Congela a primeira linha (cabeçalho) para facilitar a rolagem -----
            sheet.freeze_panes = "A2"

            wb.save(caminho)  # Salva o arquivo no caminho escolhido

            QMessageBox.information(
                self, "Sucesso",
                f"{len(resultados)} cadastro(s) exportado(s) com sucesso!\n\n{caminho}"
            )

        except Exception as e:
            QMessageBox.critical(self, "Erro", f"Falha ao exportar os cadastros:\n{str(e)}")